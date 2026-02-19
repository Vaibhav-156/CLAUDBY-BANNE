#!/usr/bin/env python3
"""
============================================
EMAIL APPROVAL MONITOR FOR CLAWED UP GLAM
============================================

This script monitors incoming approval emails from a specific email ID
and updates the text file only when a client is approved.

Features:
- Monitors incoming emails via IMAP
- Filters emails from authorized sender only
- Parses subject/body for approval/rejection keywords
- Updates text file for approved clients only (permanent storage)
- Prevents duplicate entries
- Comprehensive logging and error handling
- Secure token validation
- No auto-cleanup (manual deletion required)

Author: Automated Email Processing System
Version: 2.0.0
"""

import imaplib
import email
from email.header import decode_header
import os
import re
import json
import hashlib
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Tuple, List
import time

# Third-party imports
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from dotenv import load_dotenv
except ImportError as e:
    print(f"Missing required package: {e}")
    print("Please run: pip install -r requirements_email_monitor.txt")
    exit(1)

# ============================================
# CONFIGURATION
# ============================================

load_dotenv()

class Config:
    """Configuration settings loaded from environment variables."""
    
    # Email settings
    IMAP_SERVER = os.getenv('IMAP_SERVER', 'imap.gmail.com')
    IMAP_PORT = int(os.getenv('IMAP_PORT', '993'))
    EMAIL_ADDRESS = os.getenv('MONITOR_EMAIL')  # Email to monitor
    EMAIL_PASSWORD = os.getenv('MONITOR_EMAIL_PASSWORD')  # App password
    
    # Security - Only process emails FROM this specific email address
    AUTHORIZED_SENDER = os.getenv('AUTHORIZED_SENDER_EMAIL')
    
    # File paths
    BASE_DIR = Path(__file__).parent
    DATA_DIR = BASE_DIR / 'data'
    EXCEL_FILE = DATA_DIR / 'appointments.xlsx'  # Legacy - kept for reference
    APPROVED_CLIENTS_FILE = DATA_DIR / 'approved_clients.txt'
    PENDING_FILE = DATA_DIR / 'pending_appointments.json'
    LOG_FILE = DATA_DIR / 'email_monitor.log'
    PROCESSED_FILE = DATA_DIR / 'processed_emails.json'
    
    # Polling interval (seconds)
    POLL_INTERVAL = int(os.getenv('POLL_INTERVAL', '30'))
    
    # Approval/Rejection keywords
    APPROVAL_KEYWORDS = [
        'approved', 'approve', 'accept', 'accepted', 'confirmed', 'confirm',
        'yes', 'granted', 'ok', 'okay', 'agreed'
    ]
    
    REJECTION_KEYWORDS = [
        'declined', 'decline', 'reject', 'rejected', 'denied', 'deny',
        'no', 'refused', 'cancel', 'cancelled', 'not approved'
    ]


# ============================================
# LOGGING SETUP
# ============================================

def setup_logging():
    """Configure comprehensive logging."""
    Config.DATA_DIR.mkdir(parents=True, exist_ok=True)
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=[
            logging.FileHandler(Config.LOG_FILE, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

logger = setup_logging()


# ============================================
# EMAIL PROCESSING UTILITIES
# ============================================

def decode_email_subject(subject) -> str:
    """Decode email subject from various encodings."""
    if subject is None:
        return ""
    
    decoded_parts = decode_header(subject)
    decoded_subject = ""
    
    for part, encoding in decoded_parts:
        if isinstance(part, bytes):
            decoded_subject += part.decode(encoding or 'utf-8', errors='replace')
        else:
            decoded_subject += part
    
    return decoded_subject.strip()


def get_email_body(msg) -> str:
    """Extract text body from email message."""
    body = ""
    
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            content_disposition = str(part.get("Content-Disposition"))
            
            # Skip attachments
            if "attachment" in content_disposition:
                continue
            
            if content_type == "text/plain":
                try:
                    body = part.get_payload(decode=True).decode('utf-8', errors='replace')
                    break
                except Exception:
                    continue
    else:
        try:
            body = msg.get_payload(decode=True).decode('utf-8', errors='replace')
        except Exception:
            body = str(msg.get_payload())
    
    return body.strip()


def get_sender_email(msg) -> str:
    """Extract sender email address from message."""
    from_header = msg.get('From', '')
    
    # Extract email from format like "Name <email@domain.com>"
    match = re.search(r'<([^>]+)>', from_header)
    if match:
        return match.group(1).lower().strip()
    
    # If no angle brackets, assume the whole thing is an email
    return from_header.lower().strip()


def extract_confirmation_id(text: str) -> Optional[str]:
    """Extract confirmation ID from email text."""
    # Pattern: CUG-XXXXXXXX-XXXX
    pattern = r'CUG-[A-Z0-9]+-[A-Z0-9]+'
    match = re.search(pattern, text.upper())
    return match.group(0) if match else None


def extract_client_info_from_email(text: str) -> Dict:
    """Extract client information from email body."""
    info = {}
    
    # Common patterns for extracting info
    patterns = {
        'name': [r'(?:Name|Client):\s*([^\n]+)', r'(?:for|from)\s+([A-Z][a-z]+\s+[A-Z][a-z]+)'],
        'email': [r'(?:Email):\s*([^\s]+@[^\s]+)', r'([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)'],
        'phone': [r'(?:Phone|Tel|Mobile):\s*([0-9\-\+\(\)\s]{10,})', r'\b(\d{10})\b'],
        'date': [r'(?:Date):\s*([^\n]+)', r'(\d{4}-\d{2}-\d{2})'],
        'time': [r'(?:Time):\s*([^\n]+)', r'(\d{1,2}:\d{2}(?:\s*[AP]M)?)'],
        'service': [r'(?:Service):\s*([^\n]+)'],
        'place': [r'(?:Place|Location):\s*([^\n]+)']
    }
    
    for field, field_patterns in patterns.items():
        for pattern in field_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                info[field] = match.group(1).strip()
                break
    
    return info


# ============================================
# DECISION LOGIC
# ============================================

def determine_decision(subject: str, body: str) -> Tuple[Optional[str], str]:
    """
    Determine if email indicates approval or rejection.
    
    Returns:
        Tuple of (decision, reason) where decision is 'approved', 'declined', or None
    """
    combined_text = f"{subject} {body}".lower()
    
    # Check for rejection first (more specific)
    for keyword in Config.REJECTION_KEYWORDS:
        if keyword in combined_text:
            return ('declined', f"Found rejection keyword: '{keyword}'")
    
    # Check for approval
    for keyword in Config.APPROVAL_KEYWORDS:
        if keyword in combined_text:
            return ('approved', f"Found approval keyword: '{keyword}'")
    
    return (None, "No approval/rejection keywords found")


# ============================================
# PENDING APPOINTMENTS MANAGEMENT
# ============================================

def load_pending_appointments() -> Dict:
    """Load pending appointments from JSON file."""
    try:
        if Config.PENDING_FILE.exists():
            with open(Config.PENDING_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        logger.error(f"Error loading pending appointments: {e}")
    return {}


def save_pending_appointments(appointments: Dict):
    """Save pending appointments to JSON file."""
    try:
        Config.DATA_DIR.mkdir(parents=True, exist_ok=True)
        with open(Config.PENDING_FILE, 'w', encoding='utf-8') as f:
            json.dump(appointments, f, indent=2)
    except Exception as e:
        logger.error(f"Error saving pending appointments: {e}")


def find_appointment_by_confirmation(confirmation_id: str) -> Optional[Dict]:
    """Find appointment by confirmation ID."""
    appointments = load_pending_appointments()
    
    for token, appointment in appointments.items():
        if appointment.get('confirmationId', '').upper() == confirmation_id.upper():
            appointment['token'] = token
            return appointment
    
    return None


def update_appointment_status(token: str, status: str):
    """Update appointment status in pending file."""
    appointments = load_pending_appointments()
    
    if token in appointments:
        appointments[token]['status'] = status
        appointments[token]['updatedAt'] = datetime.now().isoformat()
        save_pending_appointments(appointments)
        logger.info(f"Updated appointment {token[:16]}... status to: {status}")


# ============================================
# TEXT FILE OPERATIONS (Permanent Storage)
# ============================================

def initialize_approved_clients_file():
    """Create text file with header if it doesn't exist."""
    Config.DATA_DIR.mkdir(parents=True, exist_ok=True)
    
    if not Config.APPROVED_CLIENTS_FILE.exists():
        header = """================================================================================
                    CLAWED UP GLAM - APPROVED CLIENTS
================================================================================
This file contains all approved client appointments.
Records are stored permanently and will not be automatically deleted.
Manual deletion is required when needed.
================================================================================

"""
        with open(Config.APPROVED_CLIENTS_FILE, 'w', encoding='utf-8') as f:
            f.write(header)
        logger.info(f"Created approved clients file: {Config.APPROVED_CLIENTS_FILE}")


def check_duplicate_in_text_file(confirmation_id: str) -> bool:
    """Check if confirmation ID already exists in text file."""
    try:
        if not Config.APPROVED_CLIENTS_FILE.exists():
            return False
        
        with open(Config.APPROVED_CLIENTS_FILE, 'r', encoding='utf-8') as f:
            content = f.read()
            return confirmation_id.upper() in content.upper()
    except Exception as e:
        logger.error(f"Error checking duplicate: {e}")
        return False


def save_approved_appointment_to_text_file(appointment: Dict) -> bool:
    """
    Save approved appointment to text file (permanent storage).
    Records are appended, never overwritten.
    Returns True if successful, False otherwise.
    """
    try:
        # Check for duplicate first
        if check_duplicate_in_text_file(appointment.get('confirmationId', '')):
            logger.warning(f"Duplicate entry detected: {appointment.get('confirmationId')}")
            return False
        
        initialize_approved_clients_file()
        
        approved_at = datetime.now().isoformat()
        separator = '-' * 80
        
        record = f"""
{separator}
APPROVED CLIENT RECORD
{separator}
Confirmation ID : {appointment.get('confirmationId', 'N/A')}
Approved At     : {approved_at}

CLIENT DETAILS:
  Name          : {appointment.get('name', 'N/A')}
  Email         : {appointment.get('email', 'N/A')}
  Phone         : {appointment.get('phone', 'N/A')}

APPOINTMENT DETAILS:
  Date          : {appointment.get('date', 'N/A')}
  Time          : {appointment.get('time', 'N/A')}
  Service       : {appointment.get('service', 'Not specified')}
  Location      : {appointment.get('place', 'Not specified')}
  Notes         : {appointment.get('notes', 'None')}

Status          : APPROVED
{separator}
"""
        
        # Append to file (does not overwrite)
        with open(Config.APPROVED_CLIENTS_FILE, 'a', encoding='utf-8') as f:
            f.write(record)
        
        logger.info(f"✅ Saved approved appointment to text file: {appointment.get('confirmationId')}")
        return True
        
    except Exception as e:
        logger.error(f"❌ Error saving to text file: {e}")
        return False


# Legacy Excel functions - kept for backwards compatibility
def initialize_excel_file():
    """Legacy: Create Excel file with headers if it doesn't exist."""
    Config.DATA_DIR.mkdir(parents=True, exist_ok=True)
    
    if not Config.EXCEL_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Appointments"
        
        # Define headers
        headers = [
            'Name', 'Email', 'Phone', 'Date', 'Time', 
            'Service', 'Notes', 'Place', 'ConfirmationID', 
            'Status', 'ApprovedAt'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='E91E63', end_color='E91E63', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        widths = [25, 30, 18, 15, 12, 20, 35, 20, 25, 15, 22]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        ws.row_dimensions[1].height = 25
        
        wb.save(Config.EXCEL_FILE)
        logger.info(f"Created Excel file: {Config.EXCEL_FILE}")


def check_duplicate_in_excel(confirmation_id: str) -> bool:
    """Legacy: Check if confirmation ID already exists in Excel."""
    try:
        if not Config.EXCEL_FILE.exists():
            return False
        
        wb = load_workbook(Config.EXCEL_FILE)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[8] and str(row[8]).upper() == confirmation_id.upper():
                return True
        
        return False
    except Exception as e:
        logger.error(f"Error checking duplicate: {e}")
        return False


def save_approved_appointment_to_excel(appointment: Dict) -> bool:
    """
    Legacy: Save approved appointment to Excel file.
    Note: This function is kept for backwards compatibility.
    The system now uses save_approved_appointment_to_text_file instead.
    Returns True if successful, False otherwise.
    """
    try:
        # Check for duplicate first
        if check_duplicate_in_excel(appointment.get('confirmationId', '')):
            logger.warning(f"Duplicate entry detected: {appointment.get('confirmationId')}")
            return False
        
        initialize_excel_file()
        
        wb = load_workbook(Config.EXCEL_FILE)
        ws = wb.active
        
        # Add new row
        new_row = [
            appointment.get('name', ''),
            appointment.get('email', ''),
            appointment.get('phone', ''),
            appointment.get('date', ''),
            appointment.get('time', ''),
            appointment.get('service', ''),
            appointment.get('notes', ''),
            appointment.get('place', ''),
            appointment.get('confirmationId', ''),
            'Approved',
            datetime.now().isoformat()
        ]
        
        ws.append(new_row)
        
        # Style the new row
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(vertical='center')
        
        wb.save(Config.EXCEL_FILE)
        logger.info(f"✅ Saved approved appointment to Excel: {appointment.get('confirmationId')}")
        return True
        
    except PermissionError:
        logger.error("❌ Excel file is currently open. Please close it and try again.")
        return False
    except Exception as e:
        logger.error(f"❌ Error saving to Excel: {e}")
        return False


# ============================================
# PROCESSED EMAILS TRACKING
# ============================================

def load_processed_emails() -> List[str]:
    """Load list of processed email message IDs."""
    try:
        if Config.PROCESSED_FILE.exists():
            with open(Config.PROCESSED_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return []


def save_processed_email(message_id: str):
    """Add message ID to processed list."""
    processed = load_processed_emails()
    if message_id not in processed:
        processed.append(message_id)
        # Keep only last 1000 entries
        processed = processed[-1000:]
        
        Config.DATA_DIR.mkdir(parents=True, exist_ok=True)
        with open(Config.PROCESSED_FILE, 'w', encoding='utf-8') as f:
            json.dump(processed, f)


def is_email_processed(message_id: str) -> bool:
    """Check if email has already been processed."""
    return message_id in load_processed_emails()


# ============================================
# SECURITY VALIDATION
# ============================================

def validate_sender(sender_email: str) -> bool:
    """Validate that sender is the authorized email address."""
    if not Config.AUTHORIZED_SENDER:
        logger.error("❌ AUTHORIZED_SENDER_EMAIL not configured!")
        return False
    
    authorized = Config.AUTHORIZED_SENDER.lower().strip()
    sender = sender_email.lower().strip()
    
    if sender == authorized:
        return True
    
    logger.warning(f"⚠️ Ignored email from unauthorized sender: {sender}")
    return False


# ============================================
# EMAIL MONITOR CLASS
# ============================================

class EmailApprovalMonitor:
    """Main class for monitoring approval emails."""
    
    def __init__(self):
        self.mail = None
        self.validate_config()
    
    def validate_config(self):
        """Validate required configuration."""
        required = {
            'MONITOR_EMAIL': Config.EMAIL_ADDRESS,
            'MONITOR_EMAIL_PASSWORD': Config.EMAIL_PASSWORD,
            'AUTHORIZED_SENDER_EMAIL': Config.AUTHORIZED_SENDER
        }
        
        missing = [k for k, v in required.items() if not v]
        
        if missing:
            logger.error(f"❌ Missing required configuration: {', '.join(missing)}")
            logger.error("Please check your .env file!")
            raise ValueError(f"Missing configuration: {missing}")
    
    def connect(self) -> bool:
        """Connect to IMAP server."""
        try:
            self.mail = imaplib.IMAP4_SSL(Config.IMAP_SERVER, Config.IMAP_PORT)
            self.mail.login(Config.EMAIL_ADDRESS, Config.EMAIL_PASSWORD)
            logger.info(f"✅ Connected to {Config.IMAP_SERVER}")
            return True
        except Exception as e:
            logger.error(f"❌ Failed to connect: {e}")
            return False
    
    def disconnect(self):
        """Disconnect from IMAP server."""
        if self.mail:
            try:
                self.mail.logout()
            except Exception:
                pass
    
    def fetch_new_emails(self) -> List[Tuple[str, email.message.Message]]:
        """Fetch new unread emails from authorized sender."""
        emails = []
        
        try:
            self.mail.select('INBOX')
            
            # Search for unread emails
            status, message_ids = self.mail.search(None, 'UNSEEN')
            
            if status != 'OK' or not message_ids[0]:
                return emails
            
            for msg_id in message_ids[0].split():
                try:
                    status, msg_data = self.mail.fetch(msg_id, '(RFC822)')
                    
                    if status == 'OK':
                        raw_email = msg_data[0][1]
                        msg = email.message_from_bytes(raw_email)
                        
                        # Create unique message identifier
                        message_id = msg.get('Message-ID', '') or f"{msg_id.decode()}-{datetime.now().timestamp()}"
                        
                        emails.append((message_id, msg))
                        
                except Exception as e:
                    logger.error(f"Error fetching email {msg_id}: {e}")
                    continue
        
        except Exception as e:
            logger.error(f"Error searching emails: {e}")
        
        return emails
    
    def process_email(self, message_id: str, msg: email.message.Message) -> Dict:
        """
        Process a single email and take appropriate action.
        
        Returns dict with processing results.
        """
        result = {
            'message_id': message_id,
            'processed': False,
            'action': None,
            'reason': None,
            'confirmation_id': None
        }
        
        # Skip if already processed
        if is_email_processed(message_id):
            result['reason'] = "Already processed"
            return result
        
        # Get sender and validate
        sender = get_sender_email(msg)
        if not validate_sender(sender):
            result['reason'] = f"Unauthorized sender: {sender}"
            save_processed_email(message_id)
            return result
        
        # Extract email content
        subject = decode_email_subject(msg.get('Subject', ''))
        body = get_email_body(msg)
        
        logger.info(f"📧 Processing email from {sender}")
        logger.info(f"   Subject: {subject}")
        
        # Determine decision
        decision, decision_reason = determine_decision(subject, body)
        
        if decision is None:
            result['reason'] = decision_reason
            logger.warning(f"⚠️ Could not determine decision: {decision_reason}")
            save_processed_email(message_id)
            return result
        
        # Extract confirmation ID
        combined_text = f"{subject} {body}"
        confirmation_id = extract_confirmation_id(combined_text)
        result['confirmation_id'] = confirmation_id
        
        if not confirmation_id:
            # Try to extract client info if no confirmation ID
            client_info = extract_client_info_from_email(combined_text)
            if client_info:
                result['client_info'] = client_info
            
            result['reason'] = "No confirmation ID found in email"
            logger.warning(f"⚠️ No confirmation ID found in email")
            save_processed_email(message_id)
            return result
        
        # Find the appointment
        appointment = find_appointment_by_confirmation(confirmation_id)
        
        if not appointment:
            result['reason'] = f"Appointment not found: {confirmation_id}"
            logger.warning(f"⚠️ Appointment not found: {confirmation_id}")
            save_processed_email(message_id)
            return result
        
        # Process based on decision
        if decision == 'approved':
            # Save to text file (permanent storage)
            if save_approved_appointment_to_text_file(appointment):
                result['action'] = 'approved_and_saved'
                result['processed'] = True
                
                # Update pending status
                if appointment.get('token'):
                    update_appointment_status(appointment['token'], 'approved')
                
                logger.info(f"✅ APPROVED: {appointment.get('name')} - {confirmation_id}")
            else:
                result['reason'] = "Failed to save to text file (possibly duplicate)"
                
        elif decision == 'declined':
            result['action'] = 'declined_no_update'
            result['processed'] = True
            
            # Update pending status but DO NOT save to Excel
            if appointment.get('token'):
                update_appointment_status(appointment['token'], 'declined')
            
            logger.info(f"❌ DECLINED: {appointment.get('name')} - {confirmation_id} (Excel NOT updated)")
        
        # Mark as processed
        save_processed_email(message_id)
        result['decision_reason'] = decision_reason
        
        return result
    
    def run_once(self) -> List[Dict]:
        """Run one iteration of email checking."""
        results = []
        
        if not self.connect():
            return results
        
        try:
            emails = self.fetch_new_emails()
            
            for message_id, msg in emails:
                result = self.process_email(message_id, msg)
                results.append(result)
        
        finally:
            self.disconnect()
        
        return results
    
    def run_continuous(self):
        """Run continuous monitoring loop."""
        logger.info("=" * 50)
        logger.info("🚀 EMAIL APPROVAL MONITOR STARTED")
        logger.info(f"📧 Monitoring: {Config.EMAIL_ADDRESS}")
        logger.info(f"🔐 Authorized sender: {Config.AUTHORIZED_SENDER}")
        logger.info(f"� Approved clients file: {Config.APPROVED_CLIENTS_FILE}")
        logger.info(f"⚠️  Auto-cleanup: DISABLED (permanent storage)")
        logger.info(f"⏱️  Poll interval: {Config.POLL_INTERVAL} seconds")
        logger.info("=" * 50)
        
        while True:
            try:
                results = self.run_once()
                
                for result in results:
                    if result.get('processed'):
                        action = result.get('action', 'unknown')
                        conf_id = result.get('confirmation_id', 'N/A')
                        logger.info(f"📋 Processed: {conf_id} -> {action}")
                
                time.sleep(Config.POLL_INTERVAL)
                
            except KeyboardInterrupt:
                logger.info("🛑 Monitor stopped by user")
                break
            except Exception as e:
                logger.error(f"❌ Error in monitoring loop: {e}")
                time.sleep(Config.POLL_INTERVAL)


# ============================================
# MAIN ENTRY POINT
# ============================================

def main():
    """Main entry point."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Email Approval Monitor for Clawed Up Glam'
    )
    parser.add_argument(
        '--once', 
        action='store_true', 
        help='Run once and exit (for testing)'
    )
    parser.add_argument(
        '--test-file',
        action='store_true',
        help='Test approved clients text file creation'
    )
    
    args = parser.parse_args()
    
    if args.test_file:
        print("Testing approved clients text file creation...")
        initialize_approved_clients_file()
        print(f"Text file created at: {Config.APPROVED_CLIENTS_FILE}")
        return
    
    try:
        monitor = EmailApprovalMonitor()
        
        if args.once:
            results = monitor.run_once()
            print(f"\nProcessed {len(results)} email(s)")
            for r in results:
                print(f"  - {r}")
        else:
            monitor.run_continuous()
            
    except ValueError as e:
        logger.error(str(e))
        print("\n⚠️  Configuration Error!")
        print("Please ensure your .env file contains:")
        print("  - MONITOR_EMAIL")
        print("  - MONITOR_EMAIL_PASSWORD")
        print("  - AUTHORIZED_SENDER_EMAIL")
        print("\nSee .env.email_monitor.example for a template.")


if __name__ == '__main__':
    main()
